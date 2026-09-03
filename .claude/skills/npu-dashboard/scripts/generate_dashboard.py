#!/usr/bin/env python3
"""Generate the PyTorch NPU test-case dashboard from a raw results workbook.

Reads ``all_testcases.xlsx`` (one sheet per module, one row per file/case record)
and produces two files:

- ``index.html`` — the dashboard shell. Contains the small aggregate ``DATA``
  object (injected between ``__DATA_BEGIN__`` / ``__DATA_END__``) and all
  rendering code, and loads the case details from a sibling file.
- ``cases.js`` — a single ``window.CASES = {...}`` assignment holding the compact
  per-case detail tree (module -> file -> [[nodeid_suffix, result], ...]),
  plus a flat ``window.FILES = [[module, file, gen, cases, status, priority,
  assignee], ...]`` list for the 测试文件 tab. This is kept OUT of the HTML so the
  shell stays small even when
  there are hundreds of thousands of cases.

Both files sit side by side and work fully offline (a ``<script src>`` tag, not a
blocked ``fetch``). The nodeid's file prefix is stripped and stored once per file;
the UI reconstructs the full nodeid as ``file + "::" + suffix``.

Usage:
    python3 generate_dashboard.py                     # default paths
    python3 generate_dashboard.py --input new.xlsx --output index.html
"""
import argparse
import json
import os
import sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip3 install openpyxl")

# Sentinel used in the raw data for a row that has no matched nodeid, i.e. a
# file-level (未泛化) record rather than a concrete test case.
UNMATCHED_NODEID = "(未匹配)"

# Execution-status keys the dashboard knows about, in display order.
STATUS_KEYS = ["passed", "failed", "skipped", "timeout", "error"]

# Header names (verbatim) the script looks for, with 0-based positional fallback
# when the header row does not match (robust to column reordering).
COLUMN_SPEC = {
    "file":   {"names": ["File"],   "fallback": 2},
    "nodeid": {"names": ["nodeid"], "fallback": 3},
    "result": {"names": ["执行结果"], "fallback": 4},
    "class":  {"names": ["Classification"], "fallback": 0},
    "num":    {"names": ["num"], "fallback": 3},
    "skip_cls":    {"names": ["skip分类"], "fallback": 5},
    "skip_reason": {"names": ["skip原因"], "fallback": 6},
}

DATA_BEGIN = "/*__DATA_BEGIN__*/"
DATA_END = "/*__DATA_END__*/"


def _find_column(headers, spec):
    for i, h in enumerate(headers):
        if h is not None and str(h).strip() in spec["names"]:
            return i
    return spec["fallback"]


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    return v if v is None else str(v).strip()


def _to_int(v):
    """Coerce a possibly-None/numeric cell to int, defaulting to 0."""
    try:
        return int(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def is_matched(nodeid):
    """True when a row represents a concrete, generalized test case."""
    return bool(nodeid) and nodeid != UNMATCHED_NODEID


def _canonical_module(name):
    """Fold the Tensor sub-sheets back onto ``Tensor`` (the legacy module key).
    The current schema splits Tensor into separate ``Tensor Operators`` and
    ``Tensor Types`` sheets; both are reported as a single ``Tensor`` module."""
    if name in ("Tensor Operators", "Tensor Types"):
        return "Tensor"
    return name


def build(path, blacklist_path=None):
    """One pass over the workbook -> (aggregate DATA dict, per-case detail tree,
    flat file list)."""
    # NOTE: not read_only — this workbook reports broken dimension metadata in
    # read-only mode (max_row=1), which would silently truncate iteration.
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = {ws.title for ws in wb.worksheets}
    if "all_files" in sheet_names and "all_testcases" in sheet_names:
        return build_two_tier(wb, blacklist_path)
    return build_legacy(wb)


def _is_blacklist_sheet(title):
    """True for a worksheet holding blacklisted/disabled cases — the current
    in-workbook ``黑名单跳过`` sheet or the legacy ``all_blacklist`` sheet."""
    t = str(title).strip()
    return "黑名单" in t or "blacklist" in t.lower()


def _blacklist_entries(header, rows, cls_to_sheet):
    """Yield ``(module, file, nodeid_suffix, result, skip_cls, skip_reason)``
    tuples from a blacklist sheet's header + rows.

    ``skip分类`` containing ``running`` (case-insensitive) — the "Running Skiped"
    entries — folds into ``skipped``; every other blacklisted case becomes the new
    ``blacklist_unsupported`` status. Both ``Classification`` and ``File`` are
    forward-filled (merged-cell convention)."""
    col_cls = _find_column(header, COLUMN_SPEC["class"])
    col_file = _find_column(header, COLUMN_SPEC["file"])
    col_nodeid = _find_column(header, COLUMN_SPEC["nodeid"])
    col_skip_cls = _find_column(header, COLUMN_SPEC["skip_cls"])
    col_skip_reason = _find_column(header, COLUMN_SPEC["skip_reason"])

    cur_cls = None
    cur_file = None
    for row in rows:
        cls = _cell(row, col_cls)
        if cls:
            cur_cls = cls
        f = _cell(row, col_file)
        if f:
            cur_file = f
        if not cur_file:
            continue
        nodeid = _cell(row, col_nodeid)
        if not nodeid:
            continue
        skip_cls = _cell(row, col_skip_cls) or ""
        skip_reason = _cell(row, col_skip_reason) or ""
        module = cls_to_sheet.get(cur_cls, cur_cls) or "Other"
        result = "skipped" if "running" in skip_cls.lower() else "blacklist_unsupported"
        suffix = nodeid[len(cur_file) + 2:] if nodeid.startswith(cur_file + "::") else nodeid
        yield (module, cur_file, suffix, result, skip_cls, skip_reason)


def load_blacklist(path, cls_to_sheet):
    """Read a legacy *separate* blacklist workbook and return its ``(module, file,
    nodeid_suffix, result, skip_cls, skip_reason)`` tuples. Only used when the
    input workbook has no in-sheet blacklist (``黑名单跳过``)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for sh in wb.worksheets:
        if _is_blacklist_sheet(sh.title):
            ws = sh
            break
    if ws is None:
        return []
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []
    return list(_blacklist_entries(header, rows, cls_to_sheet))


# Tracked generalization status, one sheet per module. Each row carries a
# ``Status`` cell like "🟢 Done" / "🔵 Todo" / "🟡 In Progress" / "⚪ Backlog";
# we keep just the text label (the coloured circle is a display emoji).
def _normalize_status(raw):
    if not raw:
        return ""
    raw = raw.strip()
    if raw and not raw[0].isascii():
        raw = raw[1:].strip()
    return raw


def load_status(path):
    """Read the tracking workbook and return a ``file path -> (status, priority,
    assignee)`` map.

    Only the module sheets are read (``README`` sheets are skipped). The ``File``
    column is located by header prefix (the header may be ``File`` or ``File(N)``
    where N is the per-module file count); status is read from ``Status`` (or the
    newer ``社区status``), priority from ``Priority``, and assignee from
    ``Assignee`` (or the newer ``author``). Each non-empty file row contributes
    its status, priority and assignee. Matching is by exact file path, so a
    dashboard file absent from this sheet simply has no tracking data."""
    wb = openpyxl.load_workbook(path, data_only=True)
    track = {}
    for ws in wb.worksheets:
        if ws.title.lower().startswith("readme"):
            continue
        header = next(ws.iter_rows(values_only=True), None)
        if header is None:
            continue
        col_file = col_status = col_priority = col_assignee = None
        for i, h in enumerate(header):
            if h is None:
                continue
            hs = str(h).strip()
            if col_file is None and hs.startswith("File"):
                col_file = i
            elif col_status is None and hs in ("Status", "社区status"):
                col_status = i
            elif col_priority is None and hs == "Priority":
                col_priority = i
            elif col_assignee is None and hs in ("Assignee", "author"):
                col_assignee = i
        if col_file is None or col_status is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            f = _cell(row, col_file)
            if not f:
                continue
            track[f] = (_normalize_status(_cell(row, col_status)),
                        _cell(row, col_priority) or "",
                        _cell(row, col_assignee) or "")
    return track


def build_two_tier(wb, blacklist_path=None):
    """New schema: a dedicated ``all_files`` sheet (one row per file, ``num`` =
    matched case count) plus an ``all_testcases`` sheet (one row per case).

    The module key is the *sheet name* (preserving the legacy grouping), not the
    ``Classification`` column: the per-module case sheets carry a finer
    ``Classification`` (e.g. the ``Tensor`` sheet holds ``Tensor`` / ``Tensor
    Operators`` / ``Tensor Types``), so we build a ``Classification -> sheet``
    map from those sheets and fold the fine values back onto their sheet."""
    all_files = set()
    all_gen_files = set()
    case_totals = Counter()
    sheets = {}
    detail = {}
    file_list = []  # [[module, file, gen(0/1), case_count], ...] for the 测试文件 tab

    # Map each Classification value to its owning sheet name (module). The
    # per-module case sheets are the source of truth here; Classification is
    # forward-filled per file within each sheet.
    cls_to_sheet = {}
    for ws in wb.worksheets:
        if ws.title in ("all_files", "all_testcases") or _is_blacklist_sheet(ws.title):
            continue
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            continue
        col_cls = _find_column(header, COLUMN_SPEC["class"])
        cur = None
        for row in rows:
            cls = _cell(row, col_cls)
            if cls:
                cur = cls
            if cur:
                cls_to_sheet[cur] = _canonical_module(ws.title)

    # ---- File-level tier: all_files (Classification forward-filled) ----
    ws = wb["all_files"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    col_cls = _find_column(header, COLUMN_SPEC["class"])
    col_file = _find_column(header, COLUMN_SPEC["file"])
    col_num = _find_column(header, COLUMN_SPEC["num"])

    file_module = {}
    file_gen = {}
    module_files = defaultdict(set)
    module_gen_files = defaultdict(set)

    cur_cls = None
    for row in rows:
        cls = _cell(row, col_cls)
        if cls:
            cur_cls = cls
        f = _cell(row, col_file)
        if not f:
            continue
        module = cls_to_sheet.get(cur_cls, cur_cls) or "Other"
        gen = _to_int(_cell(row, col_num)) > 0
        file_module[f] = module
        file_gen[f] = gen
        module_files[module].add(f)
        if gen:
            module_gen_files[module].add(f)

    all_files = set(file_module)
    all_gen_files = {f for f, g in file_gen.items() if g}

    # ---- Case-level tier: all_testcases (every row fully populated) ----
    ws = wb["all_testcases"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    col_cls = _find_column(header, COLUMN_SPEC["class"])
    col_file = _find_column(header, COLUMN_SPEC["file"])
    col_nodeid = _find_column(header, COLUMN_SPEC["nodeid"])
    col_result = _find_column(header, COLUMN_SPEC["result"])

    module_cases = defaultdict(Counter)
    module_file_cases = defaultdict(Counter)
    module_detail = defaultdict(dict)

    for row in rows:
        module = _cell(row, col_cls)
        f = _cell(row, col_file)
        nodeid = _cell(row, col_nodeid)
        result = _cell(row, col_result)
        if not nodeid:
            continue
        module = (cls_to_sheet.get(module, module) if module
                  else file_module.get(f) or "Other")
        result = (result or "error").lower()
        if result not in STATUS_KEYS:
            sys.stderr.write(f"[warn] all_testcases: unknown result "
                             f"{result!r} -> error\n")
            result = "error"
        case_totals[result] += 1
        module_cases[module][result] += 1
        module_file_cases[module][f] += 1

        suffix = nodeid[len(f) + 2:] if f and nodeid.startswith(f + "::") else nodeid
        module_detail[module].setdefault(f, []).append([suffix, result])

    # ---- Blacklist tier (optional): fold blacklisted cases into the totals ----
    # Running-skip entries land in "skipped"; everything else becomes the new
    # "blacklist_unsupported" status. Detail entries carry [suffix, result,
    # skip_cls, skip_reason] so the UI can show the skip分类 / skip原因.
    blacklist_total = 0
    # Distinct skip分类 values grouped by the result status they fold into, in
    # first-appearance order. The UI offers only the categories relevant to the
    # currently selected status — "Running Skiped" belongs to `skipped` while
    # "device not supported" / "cann not supported" belong to
    # `blacklist_unsupported` — rather than a flat list of every category.
    skip_cls_by_status = {"skipped": [], "blacklist_unsupported": []}
    skip_cls_seen = {"skipped": set(), "blacklist_unsupported": set()}
    # Blacklist source: the current schema carries it as an in-workbook
    # ``黑名单跳过`` sheet; the legacy schema keeps it in a separate workbook.
    blacklist_entries = []
    for ws in wb.worksheets:
        if _is_blacklist_sheet(ws.title):
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is not None:
                blacklist_entries = list(_blacklist_entries(header, rows, cls_to_sheet))
            break
    if not blacklist_entries and blacklist_path:
        blacklist_entries = load_blacklist(blacklist_path, cls_to_sheet)
    for module, f, suffix, result, skip_cls, skip_reason in blacklist_entries:
        blacklist_total += 1
        if skip_cls and skip_cls not in skip_cls_seen[result]:
            skip_cls_seen[result].add(skip_cls)
            skip_cls_by_status[result].append(skip_cls)
        case_totals[result] += 1
        module_cases[module][result] += 1
        module_file_cases[module][f] += 1
        module_detail[module].setdefault(f, []).append([suffix, result, skip_cls, skip_reason])

    # ---- Combine into per-module sheets ----
    for module in sorted(set(module_files) | set(module_cases)):
        files = module_files[module]
        gen_files = module_gen_files[module]
        st = module_cases[module]
        sheets[module] = {
            "files": len(files),
            "gen_files": len(gen_files),
            "na_files": len(files) - len(gen_files),
            "gen_cases": sum(st.values()),
            "passed": st["passed"],
            "failed": st["failed"],
            "skipped": st["skipped"],
            "timeout": st["timeout"],
            "error": st["error"],
            "blacklist_unsupported": st["blacklist_unsupported"],
        }
        if module_detail.get(module):
            detail[module] = module_detail[module]
        for f in sorted(files):
            file_list.append([module, f, 1 if f in gen_files else 0,
                              module_file_cases[module].get(f, 0)])

    files_total = len(all_files)
    files_gen = len(all_gen_files)
    files_na = files_total - files_gen
    cases_total = sum(case_totals.values())

    data = {
        "files_total": files_total,
        "files_gen": files_gen,
        "files_na": files_na,
        "files_gen_rate": round(files_gen / files_total * 100, 1) if files_total else 0.0,
        "cases_total": cases_total,
        "cases": {
            "passed": case_totals["passed"],
            "failed": case_totals["failed"],
            "skipped": case_totals["skipped"],
            "timeout": case_totals["timeout"],
            "error": case_totals["error"],
            "blacklist_unsupported": case_totals["blacklist_unsupported"],
        },
        "blacklist_total": blacklist_total,
        "skip_categories": skip_cls_by_status,
        "cases_pass_rate": round(case_totals["passed"] / cases_total * 100, 1) if cases_total else 0.0,
        "sheets": sheets,
    }
    return data, detail, file_list


def build_legacy(wb):
    """Legacy schema: one sheet per module (sheet name = module key); a nodeid of
    ``(未匹配)`` or empty marks a file-level (未泛化) record."""
    all_files = set()
    all_gen_files = set()
    case_totals = Counter()
    sheets = {}
    detail = {}
    file_list = []  # [[module, file, gen(0/1), case_count], ...] for the 测试文件 tab

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            continue  # empty sheet

        col_file = _find_column(header, COLUMN_SPEC["file"])
        col_nodeid = _find_column(header, COLUMN_SPEC["nodeid"])
        col_result = _find_column(header, COLUMN_SPEC["result"])

        files = set()
        gen_files = set()
        status = Counter()
        file_cases = Counter()
        module_detail = {}

        cur_file = None
        for row in rows:
            f = _cell(row, col_file)
            nodeid = _cell(row, col_nodeid)
            result = _cell(row, col_result)

            # The File column is only populated on the first row of each file's
            # case group; continuation rows leave it blank. Forward-fill it so
            # every case row still resolves to its owning file.
            if f:
                cur_file = f
            if not cur_file:
                continue  # blank leading row with no file context yet
            f = cur_file

            files.add(f)
            all_files.add(f)

            if is_matched(nodeid):
                gen_files.add(f)
                all_gen_files.add(f)
                result = (result or "error").lower()
                if result not in STATUS_KEYS:
                    # Unknown statuses are folded into "error" so totals reconcile.
                    sys.stderr.write(f"[warn] {ws.title}: unknown result "
                                     f"{result!r} -> error\n")
                    result = "error"
                status[result] += 1
                case_totals[result] += 1
                file_cases[f] += 1

                # Detail tree: strip the file prefix from the nodeid (stored once
                # per file); the UI reconstructs the full nodeid as file+"::"+suffix.
                suffix = nodeid[len(f) + 2:] if nodeid.startswith(f + "::") else nodeid
                module_detail.setdefault(f, []).append([suffix, result])

        if not files:
            continue  # sheet with a header but no data rows

        sheets[ws.title] = {
            "files": len(files),
            "gen_files": len(gen_files),
            "gen_cases": sum(status.values()),
            "passed": status["passed"],
            "failed": status["failed"],
            "skipped": status["skipped"],
            "blacklist_unsupported": 0,
            "timeout": status["timeout"],
            "error": status["error"],
            "na_files": len(files) - len(gen_files),
        }
        if module_detail:
            detail[ws.title] = module_detail
        for f in sorted(files):
            file_list.append([ws.title, f, 1 if f in gen_files else 0,
                              file_cases.get(f, 0)])

    files_total = len(all_files)
    files_gen = len(all_gen_files)
    files_na = files_total - files_gen
    cases_total = sum(case_totals.values())

    data = {
        "files_total": files_total,
        "files_gen": files_gen,
        "files_na": files_na,
        "files_gen_rate": round(files_gen / files_total * 100, 1) if files_total else 0.0,
        "cases_total": cases_total,
        "cases": {
            "passed": case_totals["passed"],
            "failed": case_totals["failed"],
            "skipped": case_totals["skipped"],
            "blacklist_unsupported": 0,
            "timeout": case_totals["timeout"],
            "error": case_totals["error"],
        },
        "blacklist_total": 0,
        "skip_categories": {"skipped": [], "blacklist_unsupported": []},
        "cases_pass_rate": round(case_totals["passed"] / cases_total * 100, 1) if cases_total else 0.0,
        "sheets": sheets,
    }
    return data, detail, file_list


def inject(html, begin, end, body):
    """Replace everything between `begin` and `end` (inclusive) with `begin+body+end`."""
    if begin not in html:
        sys.exit(f"error: marker {begin} not found in template")
    start = html.index(begin)
    stop = html.index(end, start)
    if stop < start:
        sys.exit(f"error: marker {end} not found after {begin}")
    return html[:start] + begin + body + end + html[stop + len(end):]


def main(argv=None):
    p = argparse.ArgumentParser(description="Regenerate the NPU test dashboard.")
    p.add_argument("--input", "-i", default="all_testcases.xlsx",
                   help="Input workbook (default: all_testcases.xlsx)")
    p.add_argument("--output", "-o", default="index.html",
                   help="Output/template HTML (default: index.html)")
    p.add_argument("--cases-out", "-c", default=None,
                   help="Where to write the case-detail JS (default: <output dir>/cases.js)")
    p.add_argument("--blacklist", "-b", default=None,
                   help="Blacklist workbook (default: blacklist_testcases.xlsx next to --input, if present)")
    p.add_argument("--status", "-s", default=None,
                   help="Tracking workbook (default: status_tracking.xlsx next to --input, if present); attaches Status/Priority/Assignee to each file")
    p.add_argument("--json-out", help="Optional: also write DATA to a .json file")
    args = p.parse_args(argv)

    blacklist = args.blacklist
    if blacklist is None:
        cand = os.path.join(os.path.dirname(args.input) or ".", "blacklist_testcases.xlsx")
        blacklist = cand if os.path.exists(cand) else None

    status_path = args.status
    if status_path is None:
        d = os.path.dirname(args.input) or "."
        for name in ("status_tracking.xlsx", "summary_report.xlsx"):
            cand = os.path.join(d, name)
            if os.path.exists(cand):
                status_path = cand
                break

    data, detail, file_list = build(args.input, blacklist)

    # Attach the tracked status/priority/assignee (if any) to each file list
    # entry, as 5th/6th/7th elements. Files absent from the tracking sheet carry
    # "" for each (untracked).
    track = load_status(status_path) if status_path else {}
    file_list = [[m, f, g, c] + list(track.get(f, ("", "", "")))
                 for m, f, g, c in file_list]

    # "Should Not Do" (无需泛化) files: ungeneralized files whose priority marks
    # them as not needing generalization. Fold them out of the 未泛化 bucket so the
    # file-level charts can show them as a distinct third category.
    snd_by_module = defaultdict(int)
    for m, f, g, c, status, priority, assignee in file_list:
        if g == 0 and priority == "Should Not Do":
            snd_by_module[m] += 1
    data["files_snd"] = sum(snd_by_module.values())
    data["files_na"] = data["files_total"] - data["files_gen"] - data["files_snd"]
    # 泛化率只统计 已泛化 + 未泛化（不含 无需泛化）
    gen_na = data["files_gen"] + data["files_na"]
    data["files_gen_rate"] = round(data["files_gen"] / gen_na * 100, 1) if gen_na else 0.0
    for name, sh in data["sheets"].items():
        sh["snd_files"] = snd_by_module.get(name, 0)
        sh["na_files"] = sh["files"] - sh["gen_files"] - sh["snd_files"]

    with open(args.output, "r", encoding="utf-8") as f:
        html = f.read()
    html = inject(html, DATA_BEGIN, DATA_END,
                  "\n" + json.dumps(data, ensure_ascii=False, indent=2) + "\n  ")
    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    cases_path = args.cases_out or os.path.join(os.path.dirname(args.output) or ".", "cases.js")
    with open(cases_path, "w", encoding="utf-8") as f:
        f.write("window.CASES=" +
                json.dumps(detail, ensure_ascii=False, separators=(",", ":")) + ";\n")
        f.write("window.FILES=" +
                json.dumps(file_list, ensure_ascii=False, separators=(",", ":")) + ";\n")

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # Concise verification summary
    n_files = sum(len(files) for files in detail.values())
    n_cases = sum(len(entries) for files in detail.values() for entries in files.values())
    cases_size = os.path.getsize(cases_path)
    print(f"files_total={data['files_total']}  files_gen={data['files_gen']} "
          f"({data['files_gen_rate']}%)  files_snd={data['files_snd']}  "
          f"cases_total={data['cases_total']} "
          f"pass_rate={data['cases_pass_rate']}%  blacklist={data['blacklist_total']}")
    n_tracked = sum(1 for it in file_list if it[4])
    n_priority = sum(1 for it in file_list if it[5])
    n_assignee = sum(1 for it in file_list if it[6])
    print(f"detail modules={len(detail)}  files={n_files}  cases={n_cases}  "
          f"files_list={len(file_list)} (status={n_tracked}, priority={n_priority}, "
          f"assignee={n_assignee})  -> {cases_path} ({cases_size/1024/1024:.2f} MB)")
    for name, sh in data["sheets"].items():
        print(f"  {name:14s} files={sh['files']:>3} gen={sh['gen_files']:>2} "
              f"na={sh['na_files']:>3} cases={sh['gen_cases']:>5} "
              f"P={sh['passed']:>5} F={sh['failed']:>5} S={sh['skipped']:>4} "
              f"T={sh['timeout']} E={sh['error']} B={sh['blacklist_unsupported']}")
    print(f"wrote {args.output} + {cases_path}")


if __name__ == "__main__":
    main()
